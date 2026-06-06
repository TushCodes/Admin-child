"""Excel import and export helpers for consignment records."""

import io
import logging
import re
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


def _normalize_header(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def import_from_workbook(
    file_obj,
    consignment_model=None,
    db=None,
    normalize_consignment_number=None,
    normalize_status=None,
    normalize_indian_pincode=None,
):
    """Import consignments from an uploaded workbook file-like object.

    Returns (added_count, skipped_count).
    Accepts optional `consignment_model` and `db` to allow caller to pass patched objects for testing.
    """
    if consignment_model is None:
        from app.admin.models import Consignment as consignment_model
    if db is None:
        from app.admin.models import db as db

    # use repository helpers where possible
    from app.admin.services import consignment_repo as repo

    if (
        normalize_consignment_number is None
        or normalize_status is None
        or normalize_indian_pincode is None
    ):
        from app.admin.services.logistics import (
            normalize_consignment_number as default_normalize_consignment_number,
            normalize_status as default_normalize_status,
            normalize_indian_pincode as default_normalize_indian_pincode,
        )

        normalize_consignment_number = (
            normalize_consignment_number or default_normalize_consignment_number
        )
        normalize_status = normalize_status or default_normalize_status
        normalize_indian_pincode = (
            normalize_indian_pincode or default_normalize_indian_pincode
        )

    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        raise ValueError("Excel file is empty.")

    normalized_headers = [_normalize_header(cell) for cell in header_cells]
    header_index = {name: index for index, name in enumerate(normalized_headers) if name}

    consignment_index = header_index.get("consignment_number")
    status_index = header_index.get("status")
    pickup_address_index = header_index.get("pickup_address")
    pickup_pincode_index = header_index.get("pickup_pincode")
    pickup_tag_index = header_index.get("pickup_tag")
    pickup_date_index = header_index.get("pickup_date")
    drop_address_index = header_index.get("drop_address")
    drop_pincode_index = header_index.get("drop_pincode")
    drop_tag_index = header_index.get("drop_tag")
    drop_date_index = header_index.get("drop_date")
    eta_index = header_index.get("eta")

    if None in (consignment_index, status_index):
        raise ValueError("Required headers: consignment_number, status")

    # Prefer the injected Consignment (useful for tests that patch the model)
    if consignment_model is not None:
        try:
            existing_numbers = {
                consignment_number_row[0]
                for consignment_number_row in consignment_model.query.with_entities(
                    consignment_model.consignment_number
                ).all()
            }
        except Exception:
            # If the injected Consignment cannot be queried (tests may not patch repo),
            # fall back to repository helper which uses the real model.
            existing_numbers = repo.query_existing_numbers()
    else:
        existing_numbers = repo.query_existing_numbers()
    file_seen = set()
    added_count = 0
    skipped_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        consignment_number = normalize_consignment_number(row[consignment_index])
        status = normalize_status(row[status_index])
        pickup_address = str(
            row[pickup_address_index]
            if pickup_address_index is not None and row[pickup_address_index] is not None
            else ""
        ).strip()
        pickup_pincode = normalize_indian_pincode(
            (
                row[pickup_pincode_index]
                if pickup_pincode_index is not None
                and row[pickup_pincode_index] is not None
                else ""
            ),
            "pickup_pincode",
        )
        pickup_tag = str(
            row[pickup_tag_index]
            if pickup_tag_index is not None and row[pickup_tag_index] is not None
            else ""
        ).strip()
        pickup_date = str(
            row[pickup_date_index]
            if pickup_date_index is not None and row[pickup_date_index] is not None
            else ""
        ).strip()
        drop_address = str(
            row[drop_address_index]
            if drop_address_index is not None and row[drop_address_index] is not None
            else ""
        ).strip()
        drop_pincode = normalize_indian_pincode(
            (
                row[drop_pincode_index]
                if drop_pincode_index is not None and row[drop_pincode_index] is not None
                else ""
            ),
            "drop_pincode",
        )
        drop_tag = str(
            row[drop_tag_index]
            if drop_tag_index is not None and row[drop_tag_index] is not None
            else ""
        ).strip()
        drop_date = str(
            row[drop_date_index]
            if drop_date_index is not None and row[drop_date_index] is not None
            else ""
        ).strip()
        eta = str(
            row[eta_index] if eta_index is not None and row[eta_index] is not None else ""
        ).strip()

        if consignment_number in existing_numbers or consignment_number in file_seen:
            skipped_count += 1
            continue

        consignment = consignment_model(
            consignment_number=consignment_number,
            status=status,
            pickup_address=pickup_address,
            pickup_pincode=pickup_pincode,
            pickup_tag=pickup_tag,
            pickup_date=pickup_date,
            drop_address=drop_address,
            drop_pincode=drop_pincode,
            drop_tag=drop_tag,
            drop_date=drop_date,
            eta=eta,
        )

        # Prefer the caller-provided `db` for session operations (tests may inject a fake db).
        if db is not None:
            db.session.add(consignment)
        else:
            repo.add(consignment)
        file_seen.add(consignment_number)
        existing_numbers.add(consignment_number)
        added_count += 1

    # Commit all added rows within a single transaction boundary.
    from app.admin.db.session import transaction

    with transaction(db):
        # all adds were performed above (either via injected `db.session.add` or repo.add)
        pass
    return added_count, skipped_count


def generate_import_template_bytes():
    """Return an Excel template that shows admins the accepted import columns."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import Template"

    sheet.append(
        [
            "consignment_number",
            "status",
            "pickup_address",
            "pickup_pincode",
            "pickup_tag",
            "pickup_date",
            "drop_address",
            "drop_pincode",
            "drop_tag",
            "drop_date",
        ]
    )

    sheet.append(
        [
            "CN001",
            "In Transit",
            "123 Main Street, New Delhi",
            "110017",
            "PICKUP-001",
            "2026-05-10",
            "456 Marine Drive, Mumbai",
            "400001",
            "DROP-001",
            "2026-05-12",
        ]
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_rows_to_workbook_bytes(rows):
    """Return an Excel workbook containing the supplied consignment rows."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Internal Consignments"

    sheet.append(
        [
            "consignment_number",
            "status",
            "pickup_tag",
            "drop_pincode",
            "pickup_date",
            "drop_date",
            "pickup_address",
            "drop_address",
        ]
    )

    for row in rows:
        sheet.append(
            [
                row.consignment_number,
                row.status,
                getattr(row, "pickup_tag", ""),
                row.drop_pincode,
                getattr(row, "pickup_date", ""),
                getattr(row, "drop_date", ""),
                getattr(row, "pickup_address", ""),
                getattr(row, "drop_address", ""),
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
